from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd

def escribir_indicadores(ws, datos_dict, titulo, header_fill, header_font, border, center):
    """Escribe indicadores en hoja."""
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=12)
    
    row = 3
    for key, datos in datos_dict.items():
        if 'indicadores' not in datos:
            continue
        
        zona, tipo_dia = key.rsplit('_', 1)
        ws[f'A{row}'] = f'{zona} - {tipo_dia}'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'INDICADOR'
        ws[f'B{row}'] = 'VALOR'
        for c in ['A', 'B']:
            ws[f'{c}{row}'].fill = header_fill
            ws[f'{c}{row}'].font = header_font
            ws[f'{c}{row}'].border = border
        row += 1
        
        for ind, val in datos['indicadores'].items():
            ws[f'A{row}'] = ind
            ws[f'B{row}'] = val
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            ws[f'B{row}'].alignment = center
            row += 1
        row += 2
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15

def escribir_oferta_llegadas(ws, df_ol, zona, header_fill, header_font, border, center):
    """Escribe tabla oferta/llegadas."""
    ws['A1'] = f'Oferta y Llegadas - {zona}'
    ws['A1'].font = Font(bold=True, size=12)
    
    headers = ['LADOS', 'OFERTA', '', 'LLEGADAS AUTOS', '', '', 'LLEGADAS MOTOS', '', '']
    subheaders = ['', 'AUTOS', 'MOTOS', 'TÍPICO', 'SÁBADO', 'DOMINGO', 'TÍPICO', 'SÁBADO', 'DOMINGO']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    
    for col, h in enumerate(subheaders, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    
    ws.merge_cells('B3:C3')
    ws.merge_cells('D3:F3')
    ws.merge_cells('G3:I3')
    
    row = 5
    for _, data in df_ol.iterrows():
        for col, key in enumerate(['LADOS', 'OFERTA_AUTOS', 'OFERTA_MOTOS', 'AUTOS_TIPICO', 'AUTOS_SABADO', 'AUTOS_DOMINGO', 'MOTOS_TIPICO', 'MOTOS_SABADO', 'MOTOS_DOMINGO'], 1):
            ws.cell(row=row, column=col, value=data[key]).border = border
        row += 1
    
    # Totales con fórmulas
    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    for col in range(2, 10):
        letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({letter}5:{letter}{row-1})')
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = border
    
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

def escribir_llegadas_oferta_ratio(ws, df_ratio, zona, header_fill, header_font, border, center):
    """Escribe tabla llegadas/oferta."""
    ws['A1'] = f'Llegadas / Oferta - {zona}'
    ws['A1'].font = Font(bold=True, size=12)
    
    headers = ['LADOS', 'OFERTA', '', 'LLEGADAS/OFERTA AUTOS', '', '', 'LLEGADAS/OFERTA MOTOS', '', '']
    subheaders = ['', 'AUTOS', 'MOTOS', 'TÍPICO', 'SÁBADO', 'DOMINGO', 'TÍPICO', 'SÁBADO', 'DOMINGO']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    
    for col, h in enumerate(subheaders, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    
    ws.merge_cells('B3:C3')
    ws.merge_cells('D3:F3')
    ws.merge_cells('G3:I3')
    
    row = 5
    for _, data in df_ratio.iterrows():
        for col, key in enumerate(['LADOS', 'OFERTA_AUTOS', 'OFERTA_MOTOS', 'LLEGADAS_OFERTA_AUTOS_TIPICO', 'LLEGADAS_OFERTA_AUTOS_SABADO', 'LLEGADAS_OFERTA_AUTOS_DOMINGO', 'LLEGADAS_OFERTA_MOTOS_TIPICO', 'LLEGADAS_OFERTA_MOTOS_SABADO', 'LLEGADAS_OFERTA_MOTOS_DOMINGO'], 1):
            ws.cell(row=row, column=col, value=data[key]).border = border
        row += 1
    
    # Promedios con fórmulas
    ws.cell(row=row, column=1, value='PROMEDIO').font = Font(bold=True)
    for col in range(4, 10):  # Solo promediar columnas de llegadas/oferta, no oferta
        letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=AVERAGE({letter}5:{letter}{row-1})')
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = border
    
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

def escribir_oferta_irt(ws, df_irt, zona, header_fill, header_font, border, center):
    """Escribe tabla oferta/IRT."""
    ws['A1'] = f'Oferta e Índice de Rotación Total (IRT) - {zona}'
    ws['A1'].font = Font(bold=True, size=12)
    
    headers = ['LADOS', 'OFERTA', '', 'IRT AUTOS', '', '', 'IRT MOTOS', '', '']
    subheaders = ['', 'AUTOS', 'MOTOS', 'TÍPICO', 'SÁBADO', 'DOMINGO', 'TÍPICO', 'SÁBADO', 'DOMINGO']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    
    for col, h in enumerate(subheaders, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    
    ws.merge_cells('B3:C3')
    ws.merge_cells('D3:F3')
    ws.merge_cells('G3:I3')
    
    row = 5
    for _, data in df_irt.iterrows():
        for col, key in enumerate(['LADOS', 'OFERTA_AUTOS', 'OFERTA_MOTOS', 'IRT_AUTOS_TIPICO', 'IRT_AUTOS_SABADO', 'IRT_AUTOS_DOMINGO', 'IRT_MOTOS_TIPICO', 'IRT_MOTOS_SABADO', 'IRT_MOTOS_DOMINGO'], 1):
            ws.cell(row=row, column=col, value=data[key]).border = border
        row += 1
    
    # Promedios con fórmulas (para IRT el promedio tiene sentido)
    ws.cell(row=row, column=1, value='PROMEDIO').font = Font(bold=True)
    for col in range(4, 10):  # Solo promediar columnas de IRT, no oferta
        letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=AVERAGE({letter}5:{letter}{row-1})')
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = border
    
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

def escribir_oferta_ocupacion(ws, df_oo, zona, header_fill, header_font, border, center):
    """Escribe tabla oferta/ocupación."""
    ws['A1'] = f'Oferta y Ocupación - {zona}'
    ws['A1'].font = Font(bold=True, size=12)
    
    headers = ['LADOS', 'OFERTA', '', 'OCUPACIÓN AUTOS', '', '', 'OCUPACIÓN MOTOS', '', '']
    subheaders = ['', 'AUTOS', 'MOTOS', 'TÍPICO', 'SÁBADO', 'DOMINGO', 'TÍPICO', 'SÁBADO', 'DOMINGO']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    
    for col, h in enumerate(subheaders, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center
    
    ws.merge_cells('B3:C3')
    ws.merge_cells('D3:F3')
    ws.merge_cells('G3:I3')
    
    row = 5
    for _, data in df_oo.iterrows():
        for col, key in enumerate(['LADOS', 'OFERTA_AUTOS', 'OFERTA_MOTOS', 'AUTOS_TIPICO', 'AUTOS_SABADO', 'AUTOS_DOMINGO', 'MOTOS_TIPICO', 'MOTOS_SABADO', 'MOTOS_DOMINGO'], 1):
            ws.cell(row=row, column=col, value=data[key]).border = border
        row += 1
    
    # Totales con fórmulas
    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    for col in range(2, 10):
        letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({letter}5:{letter}{row-1})')
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = border
    
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 12

def escribir_datos_ocupacion(ws, header_fill, header_font, resultados):
    """Escribe datos de ocupación por hora."""
    ws['A1'] = 'Ocupación por hora - Todos los análisis'
    ws['A1'].font = Font(bold=True, size=12)
    
    row = 3
    for tipo_est in ['via', 'parqueadero']:
        for tipo_veh in ['autos', 'motos']:
            for key, datos in resultados[tipo_est][tipo_veh].items():
                zona, tipo_dia = key.rsplit('_', 1)
                
                ws[f'A{row}'] = f'{tipo_est.upper()} - {tipo_veh.upper()} - {zona} - {tipo_dia}'
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for col, h in enumerate(['Hora', 'Ocupación', 'Entradas', 'Salidas'], 1):
                    ws.cell(row=row, column=col, value=h).fill = header_fill
                    ws.cell(row=row, column=col, value=h).font = header_font
                row += 1
                
                for hora in datos['ocupacion_por_hora'].keys():
                    ws.cell(row=row, column=1, value=f'{hora}:00')
                    ws.cell(row=row, column=2, value=round(datos['ocupacion_por_hora'][hora], 1))
                    ws.cell(row=row, column=3, value=round(datos['entradas_por_hora'][hora], 1))
                    ws.cell(row=row, column=4, value=round(datos['salidas_por_hora'][hora], 1))
                    row += 1
                row += 2

def escribir_duraciones(ws, resultados, header_fill, header_font, border, center):
    """Escribe datos de duraciones por vehículo (solo vías)."""
    ws['A1'] = 'Duraciones por Vehículo - Vías'
    ws['A1'].font = Font(bold=True, size=12)
    
    row = 3
    for tipo_veh in ['autos', 'motos']:
        for key, datos in resultados['via'][tipo_veh].items():
            if 'df_duraciones' not in datos or datos['df_duraciones'] is None:
                continue
            
            df_dur = datos['df_duraciones']
            if len(df_dur) == 0:
                continue
            
            zona, tipo_dia = key.rsplit('_', 1)
            
            ws[f'A{row}'] = f'{tipo_veh.upper()} - {zona} - {tipo_dia}'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Headers
            headers = ['PLACA', 'DIA', 'HORA_ENTRADA', 'HORA_SALIDA', 'DURACION_HORAS']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center
            row += 1
            
            # Datos
            for _, row_data in df_dur.iterrows():
                for col, key_col in enumerate(headers, 1):
                    value = row_data[key_col]
                    # Formatear timestamps como "HH:MM"
                    if key_col in ['HORA_ENTRADA', 'HORA_SALIDA']:
                        if hasattr(value, 'strftime'):
                            value = value.strftime("%H:%M")
                        else:
                            value = str(value)
                    elif key_col == 'DURACION_HORAS':
                        # Convertir decimal a formato HH:MM
                        horas = int(value)
                        minutos = int((value - horas) * 60)
                        value = f"{horas}:{minutos:02d}"
                    
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if key_col in ['HORA_ENTRADA', 'HORA_SALIDA', 'DURACION_HORAS']:
                        cell.alignment = center
                row += 1
            
            # Resumen
            ws.cell(row=row, column=1, value='TOTAL VEHÍCULOS:').font = Font(bold=True)
            ws.cell(row=row, column=2, value=len(df_dur)).font = Font(bold=True)
            row += 1
            
            ws.cell(row=row, column=1, value='DURACIÓN PROMEDIO:').font = Font(bold=True)
            dur_prom = df_dur['DURACION_HORAS'].mean()
            horas_prom = int(dur_prom)
            minutos_prom = int((dur_prom - horas_prom) * 60)
            ws.cell(row=row, column=2, value=f"{horas_prom}:{minutos_prom:02d}").font = Font(bold=True)
            row += 3
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

def escribir_datos_ocupacion_por_tipo(ws, header_fill, header_font, resultados, tipo_veh):
    """Escribe datos de ocupación por hora para un tipo de vehículo específico."""
    ws['A1'] = f'Ocupación por hora - {tipo_veh.upper()}'
    ws['A1'].font = Font(bold=True, size=12)
    
    row = 3
    for tipo_est in ['via', 'parqueadero']:
        for key, datos in resultados[tipo_est][tipo_veh].items():
            zona, tipo_dia = key.rsplit('_', 1)
            
            ws[f'A{row}'] = f'{tipo_est.upper()} - {zona} - {tipo_dia}'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for col, h in enumerate(['Hora', 'Ocupación', 'Entradas', 'Salidas'], 1):
                ws.cell(row=row, column=col, value=h).fill = header_fill
                ws.cell(row=row, column=col, value=h).font = header_font
            row += 1
            
            for hora in datos['ocupacion_por_hora'].keys():
                ws.cell(row=row, column=1, value=f'{hora}:00')
                ws.cell(row=row, column=2, value=round(datos['ocupacion_por_hora'][hora], 1))
                ws.cell(row=row, column=3, value=round(datos['entradas_por_hora'][hora], 1))
                ws.cell(row=row, column=4, value=round(datos['salidas_por_hora'][hora], 1))
                row += 1
            row += 2

def escribir_duraciones_por_tipo(ws, resultados, header_fill, header_font, border, center, tipo_veh):
    """Escribe datos de duraciones por vehículo para un tipo de vehículo específico (vías y parqueaderos)."""
    ws['A1'] = f'Duraciones por Vehículo - {tipo_veh.upper()}'
    ws['A1'].font = Font(bold=True, size=12)
    
    row = 3
    for tipo_est in ['via', 'parqueadero']:
        for key, datos in resultados[tipo_est][tipo_veh].items():
            if 'df_duraciones' not in datos or datos['df_duraciones'] is None:
                continue
            
            df_dur = datos['df_duraciones']
            if len(df_dur) == 0:
                continue
            
            zona, tipo_dia = key.rsplit('_', 1)
            
            ws[f'A{row}'] = f'{tipo_est.upper()} - {zona} - {tipo_dia}'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # Headers
            headers = ['PLACA', 'DIA', 'HORA_ENTRADA', 'HORA_SALIDA', 'DURACION_HORAS']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center
            row += 1
            
            # Datos
            for _, row_data in df_dur.iterrows():
                for col, key_col in enumerate(headers, 1):
                    value = row_data[key_col]
                    # Formatear timestamps como "HH:MM"
                    if key_col in ['HORA_ENTRADA', 'HORA_SALIDA']:
                        if hasattr(value, 'strftime'):
                            value = value.strftime("%H:%M")
                        else:
                            value = str(value)
                    elif key_col == 'DURACION_HORAS':
                        # Convertir decimal a formato HH:MM
                        horas = int(value)
                        minutos = int((value - horas) * 60)
                        value = f"{horas}:{minutos:02d}"
                    
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if key_col in ['HORA_ENTRADA', 'HORA_SALIDA', 'DURACION_HORAS']:
                        cell.alignment = center
                row += 1
            
            # Resumen
            ws.cell(row=row, column=1, value='TOTAL VEHÍCULOS:').font = Font(bold=True)
            ws.cell(row=row, column=2, value=len(df_dur)).font = Font(bold=True)
            row += 1
            
            ws.cell(row=row, column=1, value='DURACIÓN PROMEDIO:').font = Font(bold=True)
            dur_prom = df_dur['DURACION_HORAS'].mean()
            horas_prom = int(dur_prom)
            minutos_prom = int((dur_prom - horas_prom) * 60)
            ws.cell(row=row, column=2, value=f"{horas_prom}:{minutos_prom:02d}").font = Font(bold=True)
            row += 3
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15