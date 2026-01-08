from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from generar_tablas.escribir_datos import escribir_datos_ocupacion, escribir_indicadores, escribir_oferta_ocupacion, escribir_oferta_llegadas, escribir_oferta_irt, escribir_llegadas_oferta_ratio, escribir_duraciones, escribir_datos_ocupacion_por_tipo, escribir_duraciones_por_tipo, escribir_duraciones_especifico
import utils

def generar_excel(zonas,df_autos,df_motos,df_parqueaderos, resultados, output_dir):
    """Genera múltiples archivos Excel con resultados."""
    
    # Crear carpeta para los archivos Excel
    excel_dir = output_dir / 'analisis_estacionamiento'
    excel_dir.mkdir(exist_ok=True)
    
    header_fill = PatternFill(start_color=utils.COLORES['header'], end_color=utils.COLORES['header'], fill_type='solid')
    header_font = Font(bold=True, color=utils.COLORES['header_font'])
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')
    
    # 1. EXCEL DE INDICADORES
    wb_ind = Workbook()
    del wb_ind['Sheet']
    
    if resultados['via']['autos']:
        ws = wb_ind.create_sheet('IND_VIA_AUTOS')
        escribir_indicadores(ws, resultados['via']['autos'], 'INDICADORES - AUTOS EN VÍA', header_fill, header_font, border, center)
    
    if resultados['via']['motos']:
        ws = wb_ind.create_sheet('IND_VIA_MOTOS')
        escribir_indicadores(ws, resultados['via']['motos'], 'INDICADORES - MOTOS EN VÍA', header_fill, header_font, border, center)
    
    if resultados['parqueadero']['autos']:
        ws = wb_ind.create_sheet('IND_PARQ_AUTOS')
        escribir_indicadores(ws, resultados['parqueadero']['autos'], 'INDICADORES - AUTOS EN PARQUEADERO', header_fill, header_font, border, center)
    
    if resultados['parqueadero']['motos']:
        ws = wb_ind.create_sheet('IND_PARQ_MOTOS')
        escribir_indicadores(ws, resultados['parqueadero']['motos'], 'INDICADORES - MOTOS EN PARQUEADERO', header_fill, header_font, border, center)
    
    if len(wb_ind.worksheets) > 0:
        wb_ind.save(excel_dir / 'indicadores.xlsx')
    
    # 2. EXCEL DE OFERTA Y OCUPACIÓN
    if resultados['tablas_oferta_ocupacion']:
        wb_oo = Workbook()
        del wb_oo['Sheet']
        
        for zona, df_oo in resultados['tablas_oferta_ocupacion'].items():
            nombre = f'{utils.limpiar_nombre(zona)[:30]}'
            ws = wb_oo.create_sheet(nombre)
            escribir_oferta_ocupacion(ws, df_oo, zona, header_fill, header_font, border, center)
        
        wb_oo.save(excel_dir / 'oferta_ocupacion.xlsx')
    
    # 3. EXCEL DE OFERTA Y LLEGADAS
    if resultados['tablas_oferta_llegadas']:
        wb_ol = Workbook()
        del wb_ol['Sheet']
        
        for zona, df_ol in resultados['tablas_oferta_llegadas'].items():
            nombre = f'{utils.limpiar_nombre(zona)[:30]}'
            ws = wb_ol.create_sheet(nombre)
            escribir_oferta_llegadas(ws, df_ol, zona, header_fill, header_font, border, center)
        
        wb_ol.save(excel_dir / 'oferta_llegadas.xlsx')
    
    # 4. EXCEL DE IRT
    if resultados['tablas_oferta_irt']:
        wb_irt = Workbook()
        del wb_irt['Sheet']
        
        for zona, df_irt in resultados['tablas_oferta_irt'].items():
            nombre = f'{utils.limpiar_nombre(zona)[:30]}'
            ws = wb_irt.create_sheet(nombre)
            escribir_oferta_irt(ws, df_irt, zona, header_fill, header_font, border, center)
        
        wb_irt.save(excel_dir / 'irt.xlsx')
    
    # 5. EXCEL DE LLEGADAS/OFERTA
    if resultados['tablas_llegadas_oferta_ratio']:
        wb_lo = Workbook()
        del wb_lo['Sheet']
        
        for zona, df_ratio in resultados['tablas_llegadas_oferta_ratio'].items():
            nombre = f'{utils.limpiar_nombre(zona)[:30]}'
            ws = wb_lo.create_sheet(nombre)
            escribir_llegadas_oferta_ratio(ws, df_ratio, zona, header_fill, header_font, border, center)
        
        wb_lo.save(excel_dir / 'llegadas_oferta.xlsx')
    
    # 6. EXCEL DE DATOS OCUPACIÓN (con 2 hojas: autos y motos)
    wb_ocup = Workbook()
    del wb_ocup['Sheet']
    
    ws_autos = wb_ocup.create_sheet('AUTOS')
    escribir_datos_ocupacion_por_tipo(ws_autos, header_fill, header_font, resultados, 'autos')
    
    ws_motos = wb_ocup.create_sheet('MOTOS')
    escribir_datos_ocupacion_por_tipo(ws_motos, header_fill, header_font, resultados, 'motos')
    
    wb_ocup.save(excel_dir / 'datos_ocupacion.xlsx')
    
    # 7. EXCEL DE DURACIONES (con 4 hojas separadas por tipo de establecimiento y vehículo)
    wb_dur = Workbook()
    del wb_dur['Sheet']
    
    ws_dur_via_autos = wb_dur.create_sheet('DUR_VIA_AUTOS')
    escribir_duraciones_especifico(ws_dur_via_autos, resultados, header_fill, header_font, border, center, 'via', 'autos')
    
    ws_dur_via_motos = wb_dur.create_sheet('DUR_VIA_MOTOS')
    escribir_duraciones_especifico(ws_dur_via_motos, resultados, header_fill, header_font, border, center, 'via', 'motos')
    
    ws_dur_parq_autos = wb_dur.create_sheet('DUR_PARQ_AUTOS')
    escribir_duraciones_especifico(ws_dur_parq_autos, resultados, header_fill, header_font, border, center, 'parqueadero', 'autos')
    
    ws_dur_parq_motos = wb_dur.create_sheet('DUR_PARQ_MOTOS')
    escribir_duraciones_especifico(ws_dur_parq_motos, resultados, header_fill, header_font, border, center, 'parqueadero', 'motos')
    
    wb_dur.save(excel_dir / 'duraciones.xlsx')